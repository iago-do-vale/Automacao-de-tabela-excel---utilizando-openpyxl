from openpyxl import load_workbook

# função para criar aba de produto
def criar_aba(produto, arquivo):
    if produto not in arquivo.sheetnames:
        arquivo.create_sheet(produto)
        aba_criada = arquivo[produto]

        # cabeçalho
        aba_criada.cell(row=1, column=1).value = "Data"
        aba_criada.cell(row=1, column=2).value = "Produto"
        aba_criada.cell(row=1, column=3).value = "Categoria"
        aba_criada.cell(row=1, column=4).value = "Quantidade"
        aba_criada.cell(row=1, column=5).value = "Preço Unitário"
        aba_criada.cell(row=1, column=6).value = "Total"


# função para copiar dados para as abas correspondentes
def transferir_para_aba(aba_origem, linha_origem, aba_destino):
    linha_destino = aba_destino.max_row + 1

    for coluna in range(1, 7):  # agora copia todas as 6 colunas
        celula_origem = aba_origem.cell(row=linha_origem, column=coluna)
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna)

        celula_destino.value = celula_origem.value


# carregar arquivo
arquivo_vendas = load_workbook("VendasDoMes.xlsx")

# aba principal
tabela_aba = arquivo_vendas["VendasTotal"]

ultima_linha = tabela_aba.max_row

# percorrer linhas
for linha in range(2, ultima_linha + 1):
    produto = tabela_aba.cell(row=linha, column=2).value

    if not produto:
        break

    # cria aba se não existir
    criar_aba(produto, arquivo_vendas)

    # pega a aba de destino corretamente
    aba_destino = arquivo_vendas[produto]

    # transfere os dados
    transferir_para_aba(tabela_aba, linha, aba_destino)


# salvar arquivo
arquivo_vendas.save("tabela-2.xlsx")